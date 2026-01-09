from pathlib import Path
from readers import load_subject
from typing import Dict, List
from utils import (
    find_corresponding_landmarks,
    calculate_landmark_distances,
    analyse_landmark_distances,
    calculate_clockface_coordinates,
    align_prone_to_supine,
    copy_subject
)
from structures import Subject, RegistrarData
import numpy as np
import pandas as pd

# --- Configuration (matches `scripts/main.py` defaults) ---
ROOT_PATH_MRI = Path(r'U:\projects\volunteer_camri\old_data\mri_t2')
SOFT_TISSUE_ROOT = Path(r'U:\projects\dashboard\picker_points')
ANATOMICAL_JSON_BASE_ROOT = Path(r"U:\sandbox\jxu759\volunteer_seg\results")
SEGMENTATION_ROOT = Path(r'U:\sandbox\jxu759\volunteer_seg\results')

PRONE_RIBCAGE_ROOT = Path(r"U:\sandbox\jxu759\volunteer_prone_mesh")
SUPINE_RIBCAGE_ROOT = Path(r"U:\sandbox\jxu759\volunteer_seg\results\supine\rib_cage")

OUTPUT_DIR = Path("../output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_FILE_PATH_AVG = OUTPUT_DIR / "landmark_results_avg.xlsx"

# Use the same VL_IDS as main.py by default — modify here if you want different subjects
VL_IDS = [25,29,30,31,32,34,35,36,37,38]
# VL_IDS = [25]

POSITIONS = ["prone", "supine"]


def build_averaged_subjects(all_subjects_filtered: Dict[int, Subject], correspondences: Dict[int, List[List[str]]]) -> Dict[int, Subject]:
    """Create a new dict of Subject objects where each scan's registrar_data contains a single registrar
    named 'average' whose landmark coordinates are the mean of the two registrars ('anthony' & 'holly')
    using the correspondences mapping.

    The averaged landmark name uses the name from registrar 'anthony' (the first element of each pair).
    """
    averaged_subjects = {}

    for vl_id, filtered_subject in all_subjects_filtered.items():
        # copy to avoid mutating original
        subj_copy = copy_subject(filtered_subject)

        # Build mapping from anthony_name -> holly_name for this subject
        pairs = correspondences.get(vl_id, [])
        name_map = {a: b for a, b in pairs}

        for position, scan in subj_copy.scans.items():
            # Prepare averaged landmarks dict
            averaged_landmarks = {}

            reg_data_orig = filtered_subject.scans[position].registrar_data
            reg_a = reg_data_orig.get('anthony')
            reg_b = reg_data_orig.get('holly')

            if reg_a is None or reg_b is None:
                # If either registrar missing for this position skip averaging for this scan
                continue

            for name_a, name_b in pairs:
                # Verify both registrars have the named landmark for this position
                coord_a = reg_a.soft_tissue_landmarks.get(name_a)
                coord_b = reg_b.soft_tissue_landmarks.get(name_b)

                if coord_a is None or coord_b is None:
                    # skip if either missing
                    continue

                # Convert to numpy arrays and average
                try:
                    arr_a = np.asarray(coord_a, dtype=float)
                    arr_b = np.asarray(coord_b, dtype=float)
                    avg = (arr_a + arr_b) / 2.0
                    averaged_landmarks[name_a] = avg
                except Exception:
                    # if conversion fails, skip this landmark
                    continue

            # Replace registrar_data with averaged landmarks stored under both original registrar keys
            if averaged_landmarks:
                # Preserve any existing registrar data from the filtered_subject
                orig_reg = reg_data_orig
                new_reg = {}
                # copy existing registrars if they exist
                if 'anthony' in orig_reg:
                    new_reg['anthony'] = orig_reg['anthony']
                if 'holly' in orig_reg:
                    new_reg['holly'] = orig_reg['holly']

                # Add averaged landmarks under 'average'
                new_reg['average'] = RegistrarData(soft_tissue_landmarks=averaged_landmarks)
                scan.registrar_data = new_reg
            else:
                scan.registrar_data = {}

        averaged_subjects[vl_id] = subj_copy

    return averaged_subjects


if __name__ == '__main__':
    print("\n--- Running averaged-landmarks analysis script ---")

    # --- 1. Load subjects ---
    all_subjects: Dict[int, Subject] = {}
    for vl_id in VL_IDS:
        vl_id_str = f"VL{vl_id:05d}"
        print(f"--- Loading Subject: {vl_id_str} ---")
        subject = load_subject(
            vl_id=vl_id,
            positions=POSITIONS,
            dicom_root=ROOT_PATH_MRI,
            anatomical_json_base_root=ANATOMICAL_JSON_BASE_ROOT,
            soft_tissue_root=SOFT_TISSUE_ROOT
        )
        if subject.scans:
            all_subjects[vl_id] = subject

    # --- 2. Find correspondences and filtered subjects (same as main) ---
    correspondences, all_subjects_filtered = find_corresponding_landmarks(all_subjects)

    # --- 3. Build averaged-subjects dataset ---
    averaged_subjects = build_averaged_subjects(all_subjects_filtered, correspondences)

    # --- 4. Run the same analyses on averaged landmarks ---
    print("\n--- Calculating distances for averaged landmarks ---")
    distance_results_avg = calculate_landmark_distances(averaged_subjects, SEGMENTATION_ROOT)
    distance_stats_avg = analyse_landmark_distances(distance_results_avg)

    print("\n--- Averaged Landmark distance Analysis (All Subjects) ---")
    print(f"  Total Landmarks Analyzed: {distance_stats_avg.get('landmark_count', 0)}")
    print(f"    Avg. 10-Neighbor Dist Skin: {distance_stats_avg.get('mask_skin_neighborhood_avg', 0):.2f} mm")
    print(f"    Avg. 10-Neighbor Dist Rib: {distance_stats_avg.get('mask_rib_neighborhood_avg', 0):.2f} mm")

    # --- 5. Calculate clockface coordinates for averaged landmarks ---
    print("\n--- Calculating clockface coordinates for averaged landmarks ---")
    clockface_results_avg = calculate_clockface_coordinates(averaged_subjects)

    # --- 6. Alignment using averaged landmarks ---
    print("\n--- Starting Prone-to-Supine Alignment (Averaged Landmarks) ---")
    alignment_results_all_avg = {}

    for vl_id, filtered_subject in averaged_subjects.items():
        vl_id_str = filtered_subject.subject_id
        print(f"\n--- Running Alignment for {vl_id_str} ---")

        try:
            prone_mesh_file = PRONE_RIBCAGE_ROOT / f"{vl_id_str}_ribcage_prone.mesh"
            supine_seg_file = SUPINE_RIBCAGE_ROOT / f"rib_cage_{vl_id_str}.nii.gz"

            if not prone_mesh_file.exists():
                print(f"Skipping: Prone mesh not found at {prone_mesh_file}")
                continue
            if not supine_seg_file.exists():
                print(f"Skipping: Supine segmentation not found at {supine_seg_file}")
                continue

            alignment_results = align_prone_to_supine(
                subject=filtered_subject,
                prone_ribcage_mesh_path=prone_mesh_file,
                supine_ribcage_seg_path=supine_seg_file,
                orientation_flag='RAI',
                plot_for_debug=False
            )

            alignment_results_all_avg[vl_id] = alignment_results
            print(f"  Alignment for {vl_id_str} COMPLETE")
            print(f"  Ribcage Error (Mean): {alignment_results['ribcage_error_mean']:.2f} mm")

        except Exception as e:
            print(f"!!! Alignment for {vl_id_str} FAILED: {e}")

    # # --- 7. Save averaged results to Excel (separate file) ---
    # save_results_to_excel(
    #     excel_path=EXCEL_FILE_PATH_AVG,
    #     correspondences=correspondences,
    #     all_subjects=averaged_subjects,
    #     distance_results=distance_results_avg,
    #     clockface_results=clockface_results_avg,
    #     alignment_results_all=alignment_results_all_avg,
    # )

    # print(f"\nAveraged results written to: {EXCEL_FILE_PATH_AVG}")

    # --- 8. Build and save a processed sheet specifically for averaged landmarks ---
    def _get_data(results_dict: Dict, vl_id: int, position: str, key: str, lm_name: str, default=None):
        try:
            pos_dict = results_dict.get(vl_id, {}).get(position, {})
            for registrar_key in ("average", "anthony", "holly"):
                reg = pos_dict.get(registrar_key)
                if reg is None:
                    continue
                # distance-like keys are stored one level deeper under registrar_name
                if key in ("skin_distances", "rib_distances", "skin_neighborhood_avg", "rib_neighborhood_avg"):
                    val = reg.get(key, {})
                    return val.get(lm_name, default)
                else:
                    lm_entry = reg.get(lm_name)
                    if isinstance(lm_entry, dict):
                        return lm_entry.get(key, default)
            return default
        except Exception:
            return default

    def build_processed_ave_df(averaged_subjects, correspondences, distance_results, clockface_results, alignment_results):
        rows = []
        for vl_id, pairs in correspondences.items():
            if vl_id not in averaged_subjects:
                continue
            subject = averaged_subjects[vl_id]
            align_res = alignment_results.get(vl_id, {})

            nipple_disp_mag = align_res.get("nipple_displacement_magnitudes", [None, None])
            nipple_disp_vec = align_res.get("nipple_displacement_vectors", [[None] * 3, [None] * 3])

            for i, pair in enumerate(pairs):
                lm_name = pair[0]  # averaged stored under the 'anthony' name

                if align_res:
                    # alignment results may be missing r1/r2 for averaged subjects; default safely
                    lm_disp_mag = (align_res.get("r1_displacement_magnitudes", [None] * len(pairs))[i]
                                   if align_res.get("r1_displacement_magnitudes") else None)
                    lm_disp_vec = (align_res.get("r1_displacement_vectors", [[None] * 3] * len(pairs))[i]
                                  if align_res.get("r1_displacement_vectors") else [None] * 3)
                    lm_rel_mag = (align_res.get("r1_rel_nipple_magnitudes", [None] * len(pairs))[i]
                                  if align_res.get("r1_rel_nipple_magnitudes") else None)
                    lm_rel_vec = (align_res.get("r1_rel_nipple_vectors", [[None] * 3] * len(pairs))[i]
                                  if align_res.get("r1_rel_nipple_vectors") else [None] * 3)
                    ribcage_inlier_RMSE = align_res.get("ribcage_inlier_RMSE")
                else:
                    lm_disp_mag, lm_disp_vec, lm_rel_mag, lm_rel_vec, ribcage_inlier_RMSE = (
                        None, [None] * 3, None, [None] * 3, None)

                row = {
                    'Registrar': 0,  # 0 denotes averaged landmarks
                    'VL_ID': vl_id,
                    'Age': subject.age,
                    'Height [m]': subject.height,
                    'Weight [kg]': subject.weight,
                    'Landmark name': lm_name,
                    'Landmark type': lm_name.split('_')[0],

                    'landmark side (prone)': _get_data(clockface_results, vl_id, "prone", "side", lm_name),
                    'Distance to nipple (prone) [mm]': _get_data(clockface_results, vl_id, "prone", "dist_to_nipple", lm_name),
                    'Distance to rib cage (prone) [mm]': _get_data(distance_results, vl_id, "prone", "rib_distances", lm_name),
                    'Distance to skin (prone) [mm]': _get_data(distance_results, vl_id, "prone", "skin_distances", lm_name),
                    'Time (prone)': _get_data(clockface_results, vl_id, "prone", "time", lm_name),
                    'Quadrant (prone)': _get_data(clockface_results, vl_id, "prone", "quadrant", lm_name),

                    'landmark side (supine)': _get_data(clockface_results, vl_id, "supine", "side", lm_name),
                    'Distance to nipple (supine) [mm]': _get_data(clockface_results, vl_id, "supine", "dist_to_nipple", lm_name),
                    'Distance to rib cage (supine) [mm]': _get_data(distance_results, vl_id, "supine", "rib_distances", lm_name),
                    'Distance to skin (supine) [mm]': _get_data(distance_results, vl_id, "supine", "skin_distances", lm_name),
                    'Time (supine)': _get_data(clockface_results, vl_id, "supine", "time", lm_name),
                    'Quadrant (supine)': _get_data(clockface_results, vl_id, "supine", "quadrant", lm_name),

                    "ribcage_error_mean": align_res.get("ribcage_error_mean"),
                    "ribcage_error_std": align_res.get("ribcage_error_std"),
                    "ribcage_inlier_RMSE": ribcage_inlier_RMSE,
                    'Landmark displacement [mm]': lm_disp_mag,
                    'Landmark displacement relative to nipple [mm]': lm_rel_mag,
                    'Left nipple displacement [mm]': nipple_disp_mag[0],
                    'Right nipple displacement [mm]': nipple_disp_mag[1],

                    'Landmark displacement vector vx': lm_disp_vec[0] if lm_disp_vec else None,
                    'Landmark displacement vector vy': lm_disp_vec[1] if lm_disp_vec else None,
                    'Landmark displacement vector vz': lm_disp_vec[2] if lm_disp_vec else None,

                    'Landmark relative to nipple vector vx': lm_rel_vec[0] if lm_rel_vec else None,
                    'Landmark relative to nipple vector vy': lm_rel_vec[1] if lm_rel_vec else None,
                    'Landmark relative to nipple vector vz': lm_rel_vec[2] if lm_rel_vec else None,

                    'Left nipple displacement vector vx': nipple_disp_vec[0][0],
                    'Left nipple displacement vector vy': nipple_disp_vec[0][1],
                    'Left nipple displacement vector vz': nipple_disp_vec[0][2],

                    'Right nipple displacement vector vx': nipple_disp_vec[1][0],
                    'Right nipple displacement vector vy': nipple_disp_vec[1][1],
                    'Right nipple displacement vector vz': nipple_disp_vec[1][2],

                    'Mask skin neighborhood avg (prone)': _get_data(distance_results, vl_id, "prone", "skin_neighborhood_avg", lm_name),
                    'Mask rib neighborhood avg (prone)': _get_data(distance_results, vl_id, "prone", "rib_neighborhood_avg", lm_name),
                    'Mask skin neighborhood avg (supine)': _get_data(distance_results, vl_id, "supine", "skin_neighborhood_avg", lm_name),
                    'Mask rib neighborhood avg (supine)': _get_data(distance_results, vl_id, "supine", "rib_neighborhood_avg", lm_name),
                }

                rows.append(row)

        return pd.DataFrame(rows)

    try:
        df_ave = build_processed_ave_df(averaged_subjects, correspondences, distance_results_avg, clockface_results_avg, alignment_results_all_avg)

        # If the output file exists and has an existing sheet, read it and remove overlapping VL_IDs
        new_vl_ids = df_ave['VL_ID'].unique() if not df_ave.empty else []
        df_existing = pd.DataFrame()
        if EXCEL_FILE_PATH_AVG.exists():
            try:
                df_existing = pd.read_excel(EXCEL_FILE_PATH_AVG, sheet_name='processed_ave_data', engine='openpyxl')
                if not df_existing.empty:
                    df_existing_filtered = df_existing[~df_existing['VL_ID'].isin(new_vl_ids)].copy()
                    df_existing = df_existing_filtered
            except ValueError:
                # Sheet not found; we'll create it
                print(f"Warning: Sheet 'processed_ave_data' not found in {EXCEL_FILE_PATH_AVG}. Creating new sheet.")

        # Combine existing (filtered) and new, then sort
        df_combined_final = pd.concat([df_existing, df_ave], ignore_index=True) if not df_ave.empty else df_existing
        if not df_combined_final.empty:
            # Sort to keep consistent ordering similar to utils.save_results_to_excel
            sort_cols = ['Registrar', 'VL_ID']
            present_sort_cols = [c for c in sort_cols if c in df_combined_final.columns]
            if present_sort_cols:
                df_combined_final = df_combined_final.sort_values(by=present_sort_cols, kind='stable').reset_index(drop=True)

        # Write back, replacing the sheet if it exists
        try:
            # Use openpyxl load_workbook to safely replace the sheet in a file-agnostic way
            from openpyxl import load_workbook

            # If the file exists, remove the old sheet (if present) using openpyxl and save
            if EXCEL_FILE_PATH_AVG.exists():
                wb = load_workbook(EXCEL_FILE_PATH_AVG)
                if 'processed_ave_data' in wb.sheetnames:
                    std = wb['processed_ave_data']
                    wb.remove(std)
                    wb.save(EXCEL_FILE_PATH_AVG)

            # Now write the new sheet using pandas ExcelWriter in append or write mode
            with pd.ExcelWriter(EXCEL_FILE_PATH_AVG, engine='openpyxl', mode='a' if EXCEL_FILE_PATH_AVG.exists() else 'w') as writer:
                df_combined_final.to_excel(writer, sheet_name='processed_ave_data', index=False)

            print(f"Averaged processed data saved to sheet 'processed_ave_data' in {EXCEL_FILE_PATH_AVG}")
        except Exception as e:
            # Fallback: try the simple writer (may raise if pandas is too old)
            try:
                with pd.ExcelWriter(EXCEL_FILE_PATH_AVG, engine='openpyxl', mode='a' if EXCEL_FILE_PATH_AVG.exists() else 'w') as writer:
                    df_combined_final.to_excel(writer, sheet_name='processed_ave_data', index=False)
                print(f"Averaged processed data saved to sheet 'processed_ave_data' in {EXCEL_FILE_PATH_AVG}")
            except Exception as e2:
                print(f"Failed to write processed_ave_data sheet: {e} | fallback error: {e2}")
    except Exception as e:
        print(f"Failed to write processed_ave_data sheet: {e}")
