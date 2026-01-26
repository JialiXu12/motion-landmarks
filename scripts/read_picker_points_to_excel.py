"""
Script to read JSON landmark files from three folder roots and save results to Excel.

Folder roots:
- U:\projects\dashboard\picker_points\ben_original
- U:\projects\dashboard\picker_points\ben_reviewed
- U:\projects\dashboard\picker_points\holly

Each folder contains subfolders for participants, with point.XXX.json files containing
landmark positions in prone and supine positions.

Output: Excel file with one sheet per folder root.
"""

import os
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill


def read_single_json_file(json_path: Path) -> dict | None:
    """
    Read a single point JSON file and extract relevant data.

    Args:
        json_path: Path to the JSON file

    Returns:
        Dictionary with extracted data, or None if file is invalid
    """
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)

        # Extract status (include all landmarks, even rejected ones)
        status = data.get("status", "")

        # Extract subject ID
        subject = data.get("subject", "unknown")

        # Extract landmark type
        landmark_type = data.get("type", "unknown")

        # Extract prone position coordinates
        prone_x, prone_y, prone_z = None, None, None
        if "prone_point" in data:
            prone_pt = data["prone_point"].get("point", {})
            prone_x = prone_pt.get("x")
            prone_y = prone_pt.get("y")
            prone_z = prone_pt.get("z")

        # Extract supine position coordinates
        supine_x, supine_y, supine_z = None, None, None
        if "supine_point" in data:
            supine_pt = data["supine_point"].get("point", {})
            supine_x = supine_pt.get("x")
            supine_y = supine_pt.get("y")
            supine_z = supine_pt.get("z")

        return {
            "Subject": subject,
            "Landmark Type": landmark_type,
            "Filename": json_path.name,
            "Status": status,
            "Prone X": prone_x,
            "Prone Y": prone_y,
            "Prone Z": prone_z,
            "Supine X": supine_x,
            "Supine Y": supine_y,
            "Supine Z": supine_z,
        }

    except json.JSONDecodeError as e:
        print(f"  Warning: Could not parse JSON file: {json_path}. Error: {e}")
        return None
    except Exception as e:
        print(f"  Warning: Error reading file {json_path}: {e}")
        return None


def read_folder_root(folder_root: Path) -> pd.DataFrame:
    """
    Read all JSON files from a folder root (containing participant subfolders).

    Args:
        folder_root: Path to the root folder (e.g., ben_original, ben_reviewed, holly)

    Returns:
        DataFrame with all landmark data from this folder root
    """
    all_data = []

    if not folder_root.exists():
        print(f"Warning: Folder root does not exist: {folder_root}")
        return pd.DataFrame()

    # Get all participant folders (subfolders of the root)
    participant_folders = [f for f in folder_root.iterdir() if f.is_dir()]

    print(f"  Found {len(participant_folders)} participant folders in {folder_root.name}")

    for participant_folder in sorted(participant_folders):
        participant_id = participant_folder.name

        # Get all point.XXX.json files in this participant folder
        json_files = list(participant_folder.glob("point.*.json"))

        if not json_files:
            continue

        for json_file in sorted(json_files):
            result = read_single_json_file(json_file)

            if result is not None:
                # Add participant folder name for reference
                result["Participant Folder"] = participant_id
                all_data.append(result)

    # Create DataFrame
    if all_data:
        df = pd.DataFrame(all_data)
        # Reorder columns
        column_order = [
            "Subject", "Participant Folder", "Landmark Type", "Filename", "Status",
            "Prone X", "Prone Y", "Prone Z",
            "Supine X", "Supine Y", "Supine Z"
        ]
        # Only include columns that exist
        existing_cols = [c for c in column_order if c in df.columns]
        df = df[existing_cols]
        return df
    else:
        return pd.DataFrame()


def create_comparison_df(df1: pd.DataFrame, df2: pd.DataFrame, name1: str, name2: str) -> pd.DataFrame:
    """
    Create a comparison DataFrame between two registrar datasets.

    Args:
        df1: First DataFrame
        df2: Second DataFrame
        name1: Name of first registrar
        name2: Name of second registrar

    Returns:
        DataFrame with comparison results
    """
    if df1.empty or df2.empty:
        return pd.DataFrame()

    # Create a key for matching: Subject + Filename
    df1 = df1.copy()
    df2 = df2.copy()

    # Special handling for ben vs anthony comparison:
    # Only filter based on anthony's rejections (name2)
    # For other comparisons (anthony vs holly): filter both
    if name1 == 'ben' and name2 == 'anthony':
        # Only remove landmarks that anthony marked as rejected
        if 'Status' in df2.columns:
            rejected_keys = df2[df2['Status'] == 'rejected']['Subject'] + '_' + df2[df2['Status'] == 'rejected']['Filename']
            df2 = df2[df2['Status'] != 'rejected'].copy()
            # Also remove corresponding rows from ben's data
            df1['temp_key'] = df1['Subject'] + '_' + df1['Filename']
            df1 = df1[~df1['temp_key'].isin(rejected_keys)].copy()
            df1 = df1.drop(columns=['temp_key'])
    else:
        # For other comparisons, filter out rejected landmarks from both datasets
        if 'Status' in df1.columns:
            df1 = df1[df1['Status'] != 'rejected'].copy()
        if 'Status' in df2.columns:
            df2 = df2[df2['Status'] != 'rejected'].copy()

    df1['match_key'] = df1['Subject'] + '_' + df1['Filename']
    df2['match_key'] = df2['Subject'] + '_' + df2['Filename']

    # Merge on match key
    merged = pd.merge(
        df1, df2,
        on='match_key',
        how='outer',
        suffixes=(f'_{name1}', f'_{name2}')
    )

    comparison_rows = []

    for _, row in merged.iterrows():
        comp_row = {
            'Subject': row.get(f'Subject_{name1}') or row.get(f'Subject_{name2}'),
            'Filename': row.get(f'Filename_{name1}') or row.get(f'Filename_{name2}'),
        }

        # Landmark Type comparison
        lt1 = row.get(f'Landmark Type_{name1}')
        lt2 = row.get(f'Landmark Type_{name2}')
        comp_row[f'Landmark Type ({name1})'] = lt1
        comp_row[f'Landmark Type ({name2})'] = lt2
        comp_row['Landmark Type Match'] = 'Yes' if lt1 == lt2 else 'No'

        # Coordinate columns to compare
        coords = ['Prone X', 'Prone Y', 'Prone Z', 'Supine X', 'Supine Y', 'Supine Z']

        for coord in coords:
            val1 = row.get(f'{coord}_{name1}')
            val2 = row.get(f'{coord}_{name2}')
            comp_row[f'{coord} ({name1})'] = val1
            comp_row[f'{coord} ({name2})'] = val2

            # Calculate difference
            if pd.notna(val1) and pd.notna(val2):
                diff = val2 - val1
                comp_row[f'{coord} Diff'] = round(diff, 4) if diff != 0 else 0
            else:
                comp_row[f'{coord} Diff'] = None

        # Calculate Euclidean distance for supine coordinates (for anthony vs holly)
        if name1 == 'anthony' and name2 == 'holly':
            supine_x_diff = comp_row.get('Supine X Diff')
            supine_y_diff = comp_row.get('Supine Y Diff')
            supine_z_diff = comp_row.get('Supine Z Diff')

            if all(pd.notna(d) for d in [supine_x_diff, supine_y_diff, supine_z_diff]):
                euclidean_dist = (supine_x_diff**2 + supine_y_diff**2 + supine_z_diff**2)**0.5
                comp_row['Supine Euclidean Distance'] = round(euclidean_dist, 4)
            else:
                comp_row['Supine Euclidean Distance'] = None

        comparison_rows.append(comp_row)

    df_result = pd.DataFrame(comparison_rows)

    # Reorder columns to group differences at the end
    if not df_result.empty:
        # Define column order
        base_cols = ['Subject', 'Filename',
                     f'Landmark Type ({name1})', f'Landmark Type ({name2})', 'Landmark Type Match']

        coord_cols = []
        diff_cols = []

        for coord in ['Prone X', 'Prone Y', 'Prone Z', 'Supine X', 'Supine Y', 'Supine Z']:
            coord_cols.append(f'{coord} ({name1})')
            coord_cols.append(f'{coord} ({name2})')
            diff_cols.append(f'{coord} Diff')

        # Add Euclidean distance column if it exists
        if 'Supine Euclidean Distance' in df_result.columns:
            diff_cols.append('Supine Euclidean Distance')

        # Combine in desired order: base columns, then coordinate pairs, then all differences
        ordered_cols = base_cols + coord_cols + diff_cols

        # Only include columns that exist in the dataframe
        ordered_cols = [col for col in ordered_cols if col in df_result.columns]
        df_result = df_result[ordered_cols]

    return df_result


def main():
    """Main function to read JSON files and save to Excel."""

    # Define folder roots
    folder_roots = {
        "ben": Path(r"U:\projects\dashboard\picker_points\ben_original"),
        "anthony": Path(r"U:\projects\dashboard\picker_points\ben_reviewed"),
        "holly": Path(r"U:\projects\dashboard\picker_points\holly"),
    }

    # Output path
    output_dir = Path(__file__).parent.parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y_%m_%d")
    output_file = output_dir / f"picker_points_comparison_{timestamp}.xlsx"

    print("=" * 60)
    print("Reading Picker Points JSON Files")
    print("=" * 60)

    # Dictionary to store DataFrames for each sheet
    sheets_data = {}

    for sheet_name, folder_root in folder_roots.items():
        print(f"\nProcessing: {sheet_name}")
        print(f"  Path: {folder_root}")

        df = read_folder_root(folder_root)

        if not df.empty:
            sheets_data[sheet_name] = df
            print(f"  Read {len(df)} landmarks from {df['Subject'].nunique()} subjects")
        else:
            print(f"  No data found or folder does not exist")
            # Create empty dataframe with expected columns
            sheets_data[sheet_name] = pd.DataFrame(columns=[
                "Subject", "Participant Folder", "Landmark Type", "Filename", "Status",
                "Prone X", "Prone Y", "Prone Z",
                "Supine X", "Supine Y", "Supine Z"
            ])

    # Create comparison sheets
    print("\nCreating comparison sheets...")

    if 'ben' in sheets_data and 'anthony' in sheets_data:
        ben_anthony_compare = create_comparison_df(
            sheets_data['ben'], sheets_data['anthony'], 'ben', 'anthony'
        )
        if not ben_anthony_compare.empty:
            sheets_data['compare ben and anthony'] = ben_anthony_compare
            print(f"  Created 'compare ben and anthony': {len(ben_anthony_compare)} rows")

    if 'anthony' in sheets_data and 'holly' in sheets_data:
        anthony_holly_compare = create_comparison_df(
            sheets_data['anthony'], sheets_data['holly'], 'anthony', 'holly'
        )
        if not anthony_holly_compare.empty:
            sheets_data['compare anthony and holly'] = anthony_holly_compare
            print(f"  Created 'compare anthony and holly': {len(anthony_holly_compare)} rows")

            # Calculate and print Euclidean distance statistics
            if 'Supine Euclidean Distance' in anthony_holly_compare.columns:
                euclidean_distances = anthony_holly_compare['Supine Euclidean Distance'].dropna()
                if len(euclidean_distances) > 0:
                    mean_dist = euclidean_distances.mean()
                    std_dist = euclidean_distances.std()
                    print(f"\n  Supine Euclidean Distance Statistics (anthony vs holly):")
                    print(f"    Mean: {mean_dist:.4f} mm")
                    print(f"    Std Dev: {std_dist:.4f} mm")
                    print(f"    Valid measurements: {len(euclidean_distances)}")

                    # Create refined sheet with distances <= 3mm
                    refined_df = anthony_holly_compare.copy()
                    # Filter: keep rows where Euclidean distance is <= 3mm OR is NaN
                    refined_df = refined_df[
                        (refined_df['Supine Euclidean Distance'] <= 3.0) |
                        (refined_df['Supine Euclidean Distance'].isna())
                    ]

                    sheets_data['refined from anthony and holly'] = refined_df

                    # Calculate statistics for refined data
                    refined_euclidean_distances = refined_df['Supine Euclidean Distance'].dropna()
                    if len(refined_euclidean_distances) > 0:
                        refined_mean_dist = refined_euclidean_distances.mean()
                        refined_std_dist = refined_euclidean_distances.std()

                        removed_count = len(anthony_holly_compare) - len(refined_df)
                        print(f"\n  Created 'refined from anthony and holly': {len(refined_df)} rows")
                        print(f"    Removed {removed_count} rows with Euclidean distance > 3mm")
                        print(f"\n  Refined Supine Euclidean Distance Statistics (≤ 3mm):")
                        print(f"    Mean: {refined_mean_dist:.4f} mm")
                        print(f"    Std Dev: {refined_std_dist:.4f} mm")
                        print(f"    Valid measurements: {len(refined_euclidean_distances)}")

                        # Create final sheet by removing mismatched landmark types and fibroadenoma
                        final_df = refined_df.copy()

                        # Filter out rows where Landmark Type Match = 'No'
                        if 'Landmark Type Match' in final_df.columns:
                            final_df = final_df[final_df['Landmark Type Match'] != 'No']

                        # Filter out rows where landmark type is fibroadenoma (check both anthony and holly columns)
                        anthony_lt_col = 'Landmark Type (anthony)'
                        holly_lt_col = 'Landmark Type (holly)'

                        if anthony_lt_col in final_df.columns and holly_lt_col in final_df.columns:
                            final_df = final_df[
                                (final_df[anthony_lt_col] != 'fibroadenoma') &
                                (final_df[holly_lt_col] != 'fibroadenoma')
                            ]

                        sheets_data['final'] = final_df

                        # Calculate statistics for final data
                        final_euclidean_distances = final_df['Supine Euclidean Distance'].dropna()
                        if len(final_euclidean_distances) > 0:
                            final_mean_dist = final_euclidean_distances.mean()
                            final_std_dist = final_euclidean_distances.std()

                            removed_from_refined = len(refined_df) - len(final_df)
                            print(f"\n  Created 'final': {len(final_df)} rows")
                            print(f"    Removed {removed_from_refined} rows (mismatched types and fibroadenoma)")
                            print(f"\n  Final Supine Euclidean Distance Statistics:")
                            print(f"    Mean: {final_mean_dist:.4f} mm")
                            print(f"    Std Dev: {final_std_dist:.4f} mm")
                            print(f"    Valid measurements: {len(final_euclidean_distances)}")

    # Write to Excel with multiple sheets
    print(f"\n" + "=" * 60)
    print(f"Saving to: {output_file}")

    # Define red highlight fill for rejected rows
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    # Define yellow highlight fill for Euclidean distance > 3mm
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    # Define orange highlight fill for Landmark Type Match = No
    orange_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
    # Define light blue highlight fill for fibroadenoma
    lightblue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in sheets_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  Sheet '{sheet_name}': {len(df)} rows")

            # Highlight rejected rows in red
            if not df.empty and 'Status' in df.columns:
                ws = writer.sheets[sheet_name]
                status_col_idx = df.columns.get_loc('Status') + 1  # +1 because Excel is 1-indexed

                rejected_count = 0
                for row_idx, status in enumerate(df['Status'], start=2):  # start=2 to skip header row
                    if status == 'rejected':
                        rejected_count += 1
                        # Highlight the entire row
                        for col_idx in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = red_fill

                if rejected_count > 0:
                    print(f"    Highlighted {rejected_count} rejected rows in red")

            # Highlight rows with Euclidean distance > 3mm in yellow (for comparison sheets)
            if not df.empty and 'Supine Euclidean Distance' in df.columns:
                ws = writer.sheets[sheet_name]
                euclidean_col_idx = df.columns.get_loc('Supine Euclidean Distance') + 1

                high_distance_count = 0
                for row_idx, euclidean_dist in enumerate(df['Supine Euclidean Distance'], start=2):
                    if pd.notna(euclidean_dist) and euclidean_dist > 3.0:
                        high_distance_count += 1
                        # Highlight the entire row
                        for col_idx in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

                if high_distance_count > 0:
                    print(f"    Highlighted {high_distance_count} rows with Euclidean distance > 3mm in yellow")

            # Special highlighting for "refined from anthony and holly" sheet
            if sheet_name == 'refined from anthony and holly' and not df.empty:
                ws = writer.sheets[sheet_name]

                # Highlight rows where Landmark Type Match = 'No' in orange
                if 'Landmark Type Match' in df.columns:
                    match_col_idx = df.columns.get_loc('Landmark Type Match') + 1
                    mismatch_count = 0

                    for row_idx, match_val in enumerate(df['Landmark Type Match'], start=2):
                        if match_val == 'No':
                            mismatch_count += 1
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = orange_fill

                    if mismatch_count > 0:
                        print(f"    Highlighted {mismatch_count} rows with Landmark Type Match = No in orange")

                # Highlight rows where landmark type is fibroadenoma in light blue
                anthony_lt_col = 'Landmark Type (anthony)'
                holly_lt_col = 'Landmark Type (holly)'

                if anthony_lt_col in df.columns and holly_lt_col in df.columns:
                    fibroadenoma_count = 0

                    for row_idx in range(2, len(df) + 2):  # start=2 to skip header
                        df_row_idx = row_idx - 2
                        anthony_type = df.iloc[df_row_idx][anthony_lt_col]
                        holly_type = df.iloc[df_row_idx][holly_lt_col]

                        if anthony_type == 'fibroadenoma' or holly_type == 'fibroadenoma':
                            fibroadenoma_count += 1
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = lightblue_fill

                    if fibroadenoma_count > 0:
                        print(f"    Highlighted {fibroadenoma_count} rows with fibroadenoma in light blue")

    print("=" * 60)
    print("Done!")
    print(f"Output file: {output_file}")

    # Print summary statistics
    print("\n" + "=" * 60)
    print("Summary Statistics")
    print("=" * 60)

    for sheet_name, df in sheets_data.items():
        print(f"\n{sheet_name}:")
        if not df.empty:
            print(f"  Total landmarks: {len(df)}")
            print(f"  Unique subjects: {df['Subject'].nunique()}")
            if 'Landmark Type' in df.columns:
                print(f"  Landmark types: {df['Landmark Type'].value_counts().to_dict()}")
        else:
            print(f"  No data")


if __name__ == "__main__":
    main()
